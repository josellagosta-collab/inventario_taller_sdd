import csv
from datetime import datetime, time, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from .models import Material, Categoria, MovimientoInventario
from .forms import MaterialForm, TrasladoMaterialForm
from django.db.models import Q
from django.core.paginator import Paginator
from django.db.models import Count
from documentos.models import Documento
from prestamos.models import Prestamo
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from usuarios.decorators import pertenece_a_grupo
from django.http import HttpResponse
import openpyxl
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from django.conf import settings
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from incidencias.models import Incidencia
from mantenimiento.models import PlanMantenimiento
from django.db.models import F
from prestamos.models import Prestamo, Reserva
from auditoria.services import registrar_accion


MATERIAL_AUDIT_FIELDS = [
    "codigo_inventario",
    "nombre",
    "descripcion",
    "categoria",
    "subcategoria",
    "proveedor",
    "marca",
    "modelo",
    "numero_serie",
    "cantidad",
    "stock_minimo",
    "precio_compra",
    "fecha_compra",
    "garantia_hasta",
    "estado",
    "ubicacion",
    "observaciones",
]


@login_required
def lista_materiales(request):
    materiales = Material.objects.annotate(
        total_documentos=Count("documentos", distinct=True)
    ).select_related(
        "categoria",
        "ubicacion",
    ).prefetch_related(
        "reservas"
    ).order_by("nombre")

    categorias = Categoria.objects.all()

    busqueda = request.GET.get("busqueda", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    con_reserva = request.GET.get("con_reserva", "")
    stock_bajo = request.GET.get("stock_bajo", "")

    if busqueda:
        materiales = materiales.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo_inventario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(modelo__icontains=busqueda) |
            Q(numero_serie__icontains=busqueda)
        )

    if categoria_id:
        materiales = materiales.filter(categoria_id=categoria_id)

    if estado:
        materiales = materiales.filter(estado=estado)

    if con_reserva == "1":
        materiales = materiales.filter(
            reservas__estado="activa"
        ).distinct()
        
    if stock_bajo == "1":
        materiales = materiales.filter(
            cantidad__lte=F("stock_minimo")
        )

    paginator = Paginator(materiales, 10)
    numero_pagina = request.GET.get("page")
    pagina_materiales = paginator.get_page(numero_pagina)

    return render(request, "inventario/lista_materiales.html", {
        "materiales": pagina_materiales,
        "categorias": categorias,
        "busqueda": busqueda,
        "categoria_id": categoria_id,
        "estado": estado,
        "estados": Material.ESTADOS,
        "con_reserva": con_reserva,
        "stock_bajo": stock_bajo,
    })

@login_required
def detalle_material(request, material_id):
    material = get_object_or_404(
        Material.objects.select_related(
            "categoria",
            "subcategoria",
            "proveedor",
            "ubicacion",
        ).prefetch_related(
            "movimientos__usuario",
            "documentos__usuario",
            "fotografias__usuario",
            "incidencias__usuario",
            "mantenimientos__tecnico",
            "planes_mantenimiento__responsable",
            "reservas__usuario_reserva",
            "lineas_prestamo__prestamo__usuario_receptor",
            "lineas_prestamo__prestamo__profesor_responsable",
        ),
        id=material_id,
    )
    return render(request, "inventario/detalle_material.html", {
        "material": material,
        "historial_material": construir_historial_material(material),
    })


def convertir_fecha_historial(valor):
    if isinstance(valor, datetime):
        return valor

    fecha = datetime.combine(valor, time.min)

    if timezone.is_naive(fecha):
        return timezone.make_aware(fecha)

    return fecha


def construir_historial_material(material):
    historial = []

    for movimiento in material.movimientos.all():
        historial.append({
            "fecha": movimiento.fecha,
            "tipo": "Movimiento",
            "titulo": movimiento.get_tipo_display(),
            "descripcion": movimiento.descripcion or "-",
            "usuario": movimiento.usuario.username if movimiento.usuario else "-",
            "url": "",
        })

    for documento in material.documentos.all():
        historial.append({
            "fecha": documento.fecha_subida,
            "tipo": "Documento",
            "titulo": documento.nombre,
            "descripcion": documento.get_tipo_documento_display(),
            "usuario": documento.usuario.username if documento.usuario else "-",
            "url": reverse("documentos:descargar_documento", args=[documento.id]),
        })

    for fotografia in material.fotografias.all():
        historial.append({
            "fecha": fotografia.fecha_subida,
            "tipo": "Fotografía",
            "titulo": fotografia.titulo,
            "descripcion": fotografia.descripcion or "Fotografía subida.",
            "usuario": fotografia.usuario.username if fotografia.usuario else "-",
            "url": fotografia.imagen.url if fotografia.imagen else "",
        })

    for incidencia in material.incidencias.all():
        historial.append({
            "fecha": incidencia.fecha_creacion,
            "tipo": "Incidencia",
            "titulo": incidencia.titulo,
            "descripcion": (
                f"{incidencia.get_estado_display()} - "
                f"{incidencia.get_prioridad_display()}"
            ),
            "usuario": incidencia.usuario.username if incidencia.usuario else "-",
            "url": reverse("incidencias:detalle_incidencia", args=[incidencia.id]),
        })

        if incidencia.fecha_cierre:
            historial.append({
                "fecha": incidencia.fecha_cierre,
                "tipo": "Incidencia",
                "titulo": f"Cierre de incidencia: {incidencia.titulo}",
                "descripcion": incidencia.solucion or "Incidencia cerrada.",
                "usuario": incidencia.usuario.username if incidencia.usuario else "-",
                "url": reverse("incidencias:detalle_incidencia", args=[incidencia.id]),
            })

    for mantenimiento in material.mantenimientos.all():
        descripcion = (
            f"{mantenimiento.get_resultado_display()} - "
            f"{mantenimiento.descripcion}"
        )

        if mantenimiento.proxima_revision:
            descripcion = (
                f"{descripcion}. Próxima revisión: "
                f"{mantenimiento.proxima_revision}"
            )

        historial.append({
            "fecha": convertir_fecha_historial(mantenimiento.fecha),
            "tipo": "Mantenimiento",
            "titulo": mantenimiento.get_tipo_display(),
            "descripcion": descripcion,
            "usuario": mantenimiento.tecnico.username if mantenimiento.tecnico else "-",
            "url": (
                f"{reverse('mantenimiento:lista_mantenimientos')}"
                f"?busqueda={material.codigo_inventario}"
            ),
        })

    for reserva in material.reservas.all():
        historial.append({
            "fecha": convertir_fecha_historial(reserva.fecha_reserva),
            "tipo": "Reserva",
            "titulo": f"Reserva #{reserva.id}",
            "descripcion": (
                f"{reserva.get_estado_display()} - "
                f"Recogida prevista: {reserva.fecha_prevista_recogida}"
            ),
            "usuario": reserva.usuario_reserva.username,
            "url": reverse("prestamos:detalle_reserva", args=[reserva.id]),
        })

    prestamos_incluidos = set()

    for linea in material.lineas_prestamo.all():
        prestamo = linea.prestamo

        if prestamo.id in prestamos_incluidos:
            continue

        prestamos_incluidos.add(prestamo.id)
        historial.append({
            "fecha": convertir_fecha_historial(prestamo.fecha_prestamo),
            "tipo": "Préstamo",
            "titulo": f"Préstamo #{prestamo.id}",
            "descripcion": (
                f"{prestamo.get_estado_display()} - "
                f"Devolución prevista: {prestamo.fecha_prevista_devolucion}"
            ),
            "usuario": prestamo.usuario_receptor.username,
            "url": reverse("prestamos:detalle_prestamo", args=[prestamo.id]),
        })

        if prestamo.fecha_devolucion_real:
            historial.append({
                "fecha": convertir_fecha_historial(prestamo.fecha_devolucion_real),
                "tipo": "Devolución",
                "titulo": f"Devolución préstamo #{prestamo.id}",
                "descripcion": prestamo.get_estado_display(),
                "usuario": prestamo.usuario_receptor.username,
                "url": reverse("prestamos:detalle_prestamo", args=[prestamo.id]),
            })

    return sorted(
        historial,
        key=lambda evento: evento["fecha"],
        reverse=True,
    )


def obtener_valor_auditoria_material(material, campo):
    valor = getattr(material, campo)

    if campo == "estado":
        return material.get_estado_display()

    if valor is None or valor == "":
        return "-"

    return str(valor)


def construir_snapshot_auditoria_material(material):
    return {
        campo: obtener_valor_auditoria_material(material, campo)
        for campo in MATERIAL_AUDIT_FIELDS
    }


def describir_cambios_material(snapshot_anterior, material_actualizado):
    snapshot_actual = construir_snapshot_auditoria_material(material_actualizado)
    cambios = []

    for campo in MATERIAL_AUDIT_FIELDS:
        valor_anterior = snapshot_anterior[campo]
        valor_actual = snapshot_actual[campo]

        if valor_anterior == valor_actual:
            continue

        etiqueta = material_actualizado._meta.get_field(campo).verbose_name
        cambios.append(f"{etiqueta}: {valor_anterior} -> {valor_actual}")

    return cambios


@login_required
@pertenece_a_grupo("Administradores")
def crear_material(request):
    if request.method == "POST":
        form = MaterialForm(request.POST)

        if form.is_valid():
            material = form.save()
            MovimientoInventario.objects.create(
                material=material,
                tipo="alta",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion="Alta de material desde la web"
            )
            registrar_accion(
                request,
                "crear",
                "Alta de material desde la web",
                material
            )
            return redirect("inventario:detalle_material", material_id=material.id)

    else:
        form = MaterialForm()

    return render(request, "inventario/crear_material.html", {
        "form": form
    })


@login_required
@pertenece_a_grupo("Administradores")
def editar_material(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    snapshot_anterior = construir_snapshot_auditoria_material(material)

    if request.method == "POST":
        form = MaterialForm(request.POST, instance=material)

        if form.is_valid():
            material = form.save()
            cambios = describir_cambios_material(snapshot_anterior, material)

            if cambios:
                descripcion = "Material editado. Cambios: " + "; ".join(cambios)
                MovimientoInventario.objects.create(
                    material=material,
                    tipo="edicion",
                    usuario=request.user if request.user.is_authenticated else None,
                    descripcion=descripcion
                )
                registrar_accion(
                    request,
                    "editar",
                    descripcion,
                    material
                )
            return redirect("inventario:detalle_material", material_id=material.id)

    else:
        form = MaterialForm(instance=material)

    return render(request, "inventario/editar_material.html", {
        "form": form,
        "material": material
    })


@login_required
@pertenece_a_grupo("Administradores")
def trasladar_material(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    ubicacion_anterior = str(material.ubicacion) if material.ubicacion else "Sin ubicación"

    if request.method == "POST":
        form = TrasladoMaterialForm(request.POST, material=material)

        if form.is_valid():
            nueva_ubicacion = form.cleaned_data["ubicacion"]
            observaciones = form.cleaned_data.get("observaciones", "").strip()
            material.ubicacion = nueva_ubicacion
            material.save(update_fields=["ubicacion", "fecha_actualizacion"])

            descripcion = (
                "Traslado de material. "
                f"Origen: {ubicacion_anterior}. "
                f"Destino: {nueva_ubicacion}."
            )

            if observaciones:
                descripcion = f"{descripcion} Observaciones: {observaciones}"

            MovimientoInventario.objects.create(
                material=material,
                tipo="traslado",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion=descripcion,
            )
            registrar_accion(
                request,
                "editar",
                descripcion,
                material,
            )
            return redirect("inventario:detalle_material", material_id=material.id)
    else:
        form = TrasladoMaterialForm(material=material)

    return render(request, "inventario/trasladar_material.html", {
        "form": form,
        "material": material,
        "ubicacion_actual": ubicacion_anterior,
    })


@login_required
@pertenece_a_grupo("Administradores")
def retirar_material(request, material_id):
    material = get_object_or_404(Material, id=material_id)

    if request.method == "POST":
        material.estado = "retirado"
        material.save()
        MovimientoInventario.objects.create(
            material=material,
            tipo="retirada",
            usuario=request.user if request.user.is_authenticated else None,
            descripcion="Retirada lógica de material"
        )
        registrar_accion(
            request,
            "retirar",
            "Retirada lógica de material",
            material
        )
        return redirect("inventario:lista_materiales")

    return render(request, "inventario/retirar_material.html", {
        "material": material
    })


@login_required
@pertenece_a_grupo("Administradores")
def lista_movimientos(request):
    movimientos = MovimientoInventario.objects.select_related(
        "material",
        "usuario"
    ).all()

    busqueda = request.GET.get("busqueda", "")
    tipo = request.GET.get("tipo", "")

    if busqueda:
        movimientos = movimientos.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    paginator = Paginator(movimientos, 10)
    numero_pagina = request.GET.get("page")
    pagina_movimientos = paginator.get_page(numero_pagina)

    return render(request, "inventario/lista_movimientos.html", {
        "movimientos": pagina_movimientos,
        "busqueda": busqueda,
        "tipo": tipo,
        "tipos_movimiento": MovimientoInventario.TIPOS_MOVIMIENTO,
    })


@login_required
def dashboard(request):
    hoy = timezone.now().date()

    total_materiales = Material.objects.count()

    materiales_disponibles = Material.objects.filter(
        estado="disponible"
    ).count()

    materiales_prestados = Material.objects.filter(
        estado="prestado"
    ).count()

    materiales_retirados = Material.objects.filter(
        estado="retirado"
    ).count()

    materiales_stock_bajo = Material.objects.filter(
        cantidad__lte=F("stock_minimo")
    ).count()

    total_documentos = Documento.objects.count()

    prestamos_activos = Prestamo.objects.filter(
        estado="activo"
    ).count()

    prestamos_retrasados = Prestamo.objects.filter(
        estado="activo",
        fecha_prevista_devolucion__lt=hoy
    ).count()

    total_movimientos = MovimientoInventario.objects.count()

    ultimos_materiales = Material.objects.order_by("-fecha_creacion")[:5]

    ultimos_prestamos = Prestamo.objects.select_related(
        "usuario_receptor"
    ).order_by("-fecha_prestamo")[:5]

    ultimas_incidencias = Incidencia.objects.select_related(
        "material"
    ).order_by("-fecha_creacion")[:5]

    ultimos_movimientos = MovimientoInventario.objects.select_related(
        "material",
        "usuario"
    ).order_by("-fecha")[:5]

    incidencias_abiertas = Incidencia.objects.filter(
        estado="abierta"
    ).count()

    incidencias_reparacion = Incidencia.objects.filter(
        estado="en_reparacion"
    ).count()

    incidencias_cerradas = Incidencia.objects.filter(
        estado="cerrada"
    ).count()

    incidencias_criticas = Incidencia.objects.filter(
        prioridad="critica"
    ).exclude(
        estado="cerrada"
    ).count()
    
    reservas_activas = Reserva.objects.filter(estado="activa").count()

    reservas_convertidas = Reserva.objects.filter(estado="convertida").count()

    reservas_canceladas = Reserva.objects.filter(estado="cancelada").count()

    reservas_caducadas = Reserva.objects.filter(estado="caducada").count()
    
    reservas_caducadas_pendientes = Reserva.objects.filter(
        estado="activa",
        fecha_prevista_recogida__lt=hoy
    ).count()

    fecha_limite_alerta = hoy + timedelta(
        days=PlanMantenimiento.DIAS_ALERTA_REVISION
    )

    revisiones_vencidas = PlanMantenimiento.objects.filter(
        activo=True,
        proxima_revision__lt=hoy,
    ).count()

    revisiones_proximas = PlanMantenimiento.objects.filter(
        activo=True,
        proxima_revision__gte=hoy,
        proxima_revision__lte=fecha_limite_alerta,
    ).count()

    revisiones_pendientes = revisiones_vencidas + revisiones_proximas
    
    ultimas_reservas = Reserva.objects.select_related(
        "usuario_reserva",
        "material"
    ).order_by("-fecha_reserva")[:5]

    return render(request, "inventario/dashboard.html", {
        "total_materiales": total_materiales,
        "materiales_disponibles": materiales_disponibles,
        "materiales_prestados": materiales_prestados,
        "materiales_retirados": materiales_retirados,
        "materiales_stock_bajo": materiales_stock_bajo,
        "total_documentos": total_documentos,
        "prestamos_activos": prestamos_activos,
        "prestamos_retrasados": prestamos_retrasados,
        "total_movimientos": total_movimientos,
        "ultimos_materiales": ultimos_materiales,
        "ultimos_prestamos": ultimos_prestamos,
        "ultimas_incidencias": ultimas_incidencias,
        "ultimos_movimientos": ultimos_movimientos,
        "incidencias_abiertas": incidencias_abiertas,
        "incidencias_reparacion": incidencias_reparacion,
        "incidencias_cerradas": incidencias_cerradas,
        "incidencias_criticas": incidencias_criticas,
        "reservas_activas": reservas_activas,
        "reservas_convertidas": reservas_convertidas,
        "reservas_canceladas": reservas_canceladas,
        "reservas_caducadas": reservas_caducadas,
        "reservas_caducadas_pendientes": reservas_caducadas_pendientes,
        "revisiones_vencidas": revisiones_vencidas,
        "revisiones_proximas": revisiones_proximas,
        "revisiones_pendientes": revisiones_pendientes,
        "ultimas_reservas": ultimas_reservas,
    })


@login_required
@pertenece_a_grupo("Administradores")
def exportar_materiales_csv(request):
    materiales = Material.objects.select_related(
        "categoria",
        "subcategoria",
        "proveedor",
        "ubicacion",
    ).all()

    busqueda = request.GET.get("busqueda", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    con_reserva = request.GET.get("con_reserva", "")
    stock_bajo = request.GET.get("stock_bajo", "")

    if busqueda:
        materiales = materiales.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo_inventario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(modelo__icontains=busqueda) |
            Q(numero_serie__icontains=busqueda)
        )

    if categoria_id:
        materiales = materiales.filter(categoria_id=categoria_id)

    if estado:
        materiales = materiales.filter(estado=estado)

    if con_reserva == "1":
        materiales = materiales.filter(reservas__estado="activa").distinct()

    if stock_bajo == "1":
        materiales = materiales.filter(cantidad__lte=F("stock_minimo"))

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="inventario.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "Código",
        "Nombre",
        "Categoría",
        "Marca",
        "Modelo",
        "Cantidad",
        "Estado",
    ])

    for material in materiales:
        writer.writerow([
            material.codigo_inventario,
            material.nombre,
            material.categoria.nombre if material.categoria else "",
            material.marca,
            material.modelo,
            material.cantidad,
            material.get_estado_display(),
        ])

    return response


@login_required
@pertenece_a_grupo("Administradores")
def exportar_materiales_excel(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Materiales"

    encabezados = [
        "Código",
        "Nombre",
        "Categoría",
        "Marca",
        "Modelo",
        "Cantidad",
        "Estado",
    ]

    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=texto)
        
    color_corporativo = "0051A0"

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color=color_corporativo,
            end_color=color_corporativo,
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    materiales = Material.objects.select_related(
        "categoria",
        "subcategoria",
        "proveedor",
        "ubicacion"
    ).all()

    busqueda = request.GET.get("busqueda", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    con_reserva = request.GET.get("con_reserva", "")
    stock_bajo = request.GET.get("stock_bajo", "")
   

    if busqueda:
        materiales = materiales.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo_inventario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(modelo__icontains=busqueda) |
            Q(numero_serie__icontains=busqueda)
        )

    if categoria_id:
        materiales = materiales.filter(categoria_id=categoria_id)

    if estado:
        materiales = materiales.filter(estado=estado)
        
    if con_reserva == "1":
        materiales = materiales.filter(
            reservas__estado="activa"
        ).distinct()
        
    if stock_bajo == "1":
        materiales = materiales.filter(
            cantidad__lte=F("stock_minimo")
        )

    fila = 2

    for material in materiales:
        hoja.cell(fila, 1, material.codigo_inventario)
        hoja.cell(fila, 2, material.nombre)

        hoja.cell(
            fila,
            3,
            material.categoria.nombre if material.categoria else ""
        )

        hoja.cell(fila, 4, material.marca)
        hoja.cell(fila, 5, material.modelo)
        hoja.cell(fila, 6, material.cantidad)
        hoja.cell(fila, 7, material.get_estado_display())

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'
    
    for columna in hoja.columns:
        max_length = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        hoja.column_dimensions[letra_columna].width = max_length + 2

        hoja.auto_filter.ref = hoja.dimensions
        hoja.freeze_panes = "A2"

    workbook.save(response)

    return response


@login_required
@pertenece_a_grupo("Administradores")
def exportar_materiales_pdf(request):

    buffer = BytesIO()

    pdf = SimpleDocTemplate(buffer)

    elementos = []

    estilos = getSampleStyleSheet()

    logo_path = settings.BASE_DIR / "static" / "images" / "logo_monlau.png"

    if logo_path.exists():
        logo = Image(str(logo_path), width=90, height=55)
        elementos.append(logo)

    titulo = Paragraph(
        "<b>Inventario de Taller</b>",
        estilos["Title"]
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 12))

    fecha = timezone.now().strftime("%d/%m/%Y %H:%M")

    elementos.append(
        Paragraph(
            f"Fecha de generación: {fecha}",
            estilos["Normal"]
        )
    )

    elementos.append(Spacer(1, 12))

    total_materiales = Material.objects.count()
    disponibles = Material.objects.filter(estado="disponible").count()
    prestados = Material.objects.filter(estado="prestado").count()
    averiados = Material.objects.filter(estado="averiado").count()

    resumen = Paragraph(
        f"""
        <b>Total materiales:</b> {total_materiales}<br/>
        <b>Disponibles:</b> {disponibles}<br/>
        <b>Prestados:</b> {prestados}<br/>
        <b>Averiados:</b> {averiados}
        """,
        estilos["Normal"]
    )

    elementos.append(resumen)
    elementos.append(Spacer(1, 18))

    datos = [
        [
            "Código",
            "Nombre",
            "Categoría",
            "Cantidad",
            "Estado",
        ]
    ]

    materiales = Material.objects.select_related("categoria").all()

    busqueda = request.GET.get("busqueda", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
    con_reserva = request.GET.get("con_reserva", "")
    stock_bajo = request.GET.get("stock_bajo", "")

    if busqueda:
        materiales = materiales.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo_inventario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(modelo__icontains=busqueda) |
            Q(numero_serie__icontains=busqueda)
        )

    if categoria_id:
        materiales = materiales.filter(categoria_id=categoria_id)

    if estado:
        materiales = materiales.filter(estado=estado)

    if con_reserva == "1":
        materiales = materiales.filter(
            reservas__estado="activa"
        ).distinct()
        
    if stock_bajo == "1":
        materiales = materiales.filter(
            cantidad__lte=F("stock_minimo")
        )

    for material in materiales:
        datos.append([
            material.codigo_inventario,
            material.nombre,
            material.categoria.nombre if material.categoria else "",
            material.cantidad,
            material.get_estado_display(),
        ])

    tabla = Table(datos, repeatRows=1)

    tabla.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051A0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ])
    )

    elementos.append(tabla)

    pdf.build(elementos)

    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/pdf"
    )

    response["Content-Disposition"] = 'attachment; filename="inventario_monlau.pdf"'

    return response


@login_required
@pertenece_a_grupo("Administradores")
def exportar_movimientos_csv(request):
    movimientos = MovimientoInventario.objects.select_related(
        "material",
        "usuario",
    ).all()

    busqueda = request.GET.get("busqueda", "")
    tipo = request.GET.get("tipo", "")

    if busqueda:
        movimientos = movimientos.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="movimientos.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "ID",
        "Material",
        "Código material",
        "Tipo",
        "Usuario",
        "Descripción",
        "Fecha",
    ])

    for movimiento in movimientos:
        writer.writerow([
            movimiento.id,
            movimiento.material.nombre,
            movimiento.material.codigo_inventario,
            movimiento.get_tipo_display(),
            movimiento.usuario.username if movimiento.usuario else "",
            movimiento.descripcion or "",
            movimiento.fecha.strftime("%d/%m/%Y %H:%M"),
        ])

    return response


@login_required
@pertenece_a_grupo("Administradores")
def exportar_movimientos_excel(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Movimientos"

    encabezados = [
        "ID",
        "Material",
        "Código material",
        "Tipo",
        "Usuario",
        "Descripción",
        "Fecha",
    ]

    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=texto)
        
    color_corporativo = "0051A0"

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color=color_corporativo,
            end_color=color_corporativo,
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    movimientos = MovimientoInventario.objects.select_related(
        "material",
        "usuario"
    ).all()
    
    busqueda = request.GET.get("busqueda", "")
    tipo = request.GET.get("tipo", "")

    if busqueda:
        movimientos = movimientos.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    fila = 2

    for movimiento in movimientos:
        hoja.cell(fila, 1, movimiento.id)
        hoja.cell(fila, 2, movimiento.material.nombre)
        hoja.cell(fila, 3, movimiento.material.codigo_inventario)
        hoja.cell(fila, 4, movimiento.get_tipo_display())

        if movimiento.usuario:
            hoja.cell(fila, 5, movimiento.usuario.username)
        else:
            hoja.cell(fila, 5, "")

        hoja.cell(fila, 6, movimiento.descripcion or "")

        hoja.cell(
            fila,
            7,
            movimiento.fecha.strftime("%d/%m/%Y %H:%M")
        )

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="movimientos.xlsx"'
    
    for columna in hoja.columns:
        max_length = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        hoja.column_dimensions[letra_columna].width = max_length + 2

    hoja.auto_filter.ref = hoja.dimensions
    hoja.freeze_panes = "A2"

    workbook.save(response)

    return response

@login_required
@pertenece_a_grupo("Administradores")
def exportar_movimientos_pdf(request):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer)

    elementos = []
    estilos = getSampleStyleSheet()

    logo_path = settings.BASE_DIR / "static" / "images" / "logo_monlau.png"

    if logo_path.exists():
        elementos.append(Image(str(logo_path), width=90, height=55))

    elementos.append(Paragraph("<b>Informe de movimientos</b>", estilos["Title"]))
    elementos.append(Spacer(1, 12))

    datos = [
        ["Fecha", "Material", "Tipo", "Usuario", "Descripción"]
    ]

    movimientos = MovimientoInventario.objects.select_related(
        "material",
        "usuario"
    ).all()

    busqueda = request.GET.get("busqueda", "")
    tipo = request.GET.get("tipo", "")

    if busqueda:
        movimientos = movimientos.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )

    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    for movimiento in movimientos:
        datos.append([
            movimiento.fecha.strftime("%d/%m/%Y %H:%M"),
            movimiento.material.nombre,
            movimiento.get_tipo_display(),
            movimiento.usuario.username if movimiento.usuario else "",
            movimiento.descripcion or "",
        ])

    tabla = Table(datos, repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051A0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="movimientos.pdf"'

    return response


@login_required
@pertenece_a_grupo("Administradores")
def materiales_stock_bajo(request):

    materiales = Material.objects.filter(
        cantidad__lte=F("stock_minimo")
    )

    return render(
        request,
        "inventario/materiales_stock_bajo.html",
        {
            "materiales": materiales,
        }
    )
