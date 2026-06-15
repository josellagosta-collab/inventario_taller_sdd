import csv

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Incidencia
from inventario.models import Material, MovimientoInventario
from .forms import IncidenciaForm
from django.utils import timezone
from .forms import ResolverIncidenciaForm
from django.http import HttpResponse
import openpyxl
from django.db.models import Q
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.conf import settings
from usuarios.decorators import pertenece_a_grupo

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

@login_required
@pertenece_a_grupo("Administradores")
def crear_incidencia(request, material_id):
    material = get_object_or_404(Material, id=material_id)

    if request.method == "POST":
        form = IncidenciaForm(request.POST)

        if form.is_valid():
            incidencia = form.save(commit=False)
            incidencia.material = material

            if request.user.is_authenticated:
                incidencia.usuario = request.user

            incidencia.save()

            material.estado = "averiado"
            material.save()

            MovimientoInventario.objects.create(
                material=material,
                tipo="edicion",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion=f"Incidencia creada: {incidencia.titulo}"
            )

            return redirect("inventario:detalle_material", material_id=material.id)

    else:
        form = IncidenciaForm()

    return render(request, "incidencias/crear_incidencia.html", {
        "form": form,
        "material": material,
    })


@login_required
def lista_incidencias(request):
    incidencias = Incidencia.objects.select_related(
        "material",
        "usuario"
    ).all()

    busqueda = request.GET.get("busqueda", "")
    prioridad = request.GET.get("prioridad", "")
    estado = request.GET.get("estado", "")

    if busqueda:
        incidencias = incidencias.filter(
            Q(titulo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if prioridad:
        incidencias = incidencias.filter(prioridad=prioridad)

    if estado:
        incidencias = incidencias.filter(estado=estado)

    return render(request, "incidencias/lista_incidencias.html", {
        "incidencias": incidencias,
        "busqueda": busqueda,
        "prioridad": prioridad,
        "estado": estado,
        "prioridades": Incidencia.PRIORIDADES,
        "estados": Incidencia.ESTADOS,
    })


@login_required
@pertenece_a_grupo("Administradores")
def resolver_incidencia(request, incidencia_id):

    incidencia = get_object_or_404(
        Incidencia,
        id=incidencia_id
    )

    if request.method == "POST":

        form = ResolverIncidenciaForm(request.POST)

        if form.is_valid():

            incidencia.solucion = form.cleaned_data["solucion"]

            incidencia.estado = "cerrada"

            incidencia.fecha_cierre = timezone.now()

            incidencia.save()

            material = incidencia.material

            material.estado = "disponible"

            material.save()

            MovimientoInventario.objects.create(
                material=material,
                tipo="edicion",
                usuario=request.user,
                descripcion=f"Incidencia resuelta: {incidencia.titulo}"
            )

            return redirect(
                "incidencias:detalle_incidencia",
                incidencia.id
            )

    else:

        form = ResolverIncidenciaForm()

    return render(
        request,
        "incidencias/resolver_incidencia.html",
        {
            "incidencia": incidencia,
            "form": form,
        }
    )


@login_required
def detalle_incidencia(request, incidencia_id):
    incidencia = get_object_or_404(
        Incidencia.objects.select_related(
            "material",
            "usuario"
        ).prefetch_related(
            "comentarios__usuario"
        ),
        id=incidencia_id
    )

    return render(request, "incidencias/detalle_incidencia.html", {
        "incidencia": incidencia,
    })

@login_required
@pertenece_a_grupo("Administradores")
def exportar_incidencias_csv(request):
    incidencias = Incidencia.objects.select_related(
        "material",
        "usuario",
    ).all()

    busqueda = request.GET.get("busqueda", "")
    prioridad = request.GET.get("prioridad", "")
    estado = request.GET.get("estado", "")

    if busqueda:
        incidencias = incidencias.filter(
            Q(titulo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if prioridad:
        incidencias = incidencias.filter(prioridad=prioridad)

    if estado:
        incidencias = incidencias.filter(estado=estado)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="incidencias.csv"'
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "ID",
        "Título",
        "Material",
        "Código material",
        "Prioridad",
        "Estado",
        "Usuario",
        "Fecha creación",
        "Fecha cierre",
        "Solución",
    ])

    for incidencia in incidencias:
        writer.writerow([
            incidencia.id,
            incidencia.titulo,
            incidencia.material.nombre,
            incidencia.material.codigo_inventario,
            incidencia.get_prioridad_display(),
            incidencia.get_estado_display(),
            incidencia.usuario.username if incidencia.usuario else "",
            incidencia.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            (
                incidencia.fecha_cierre.strftime("%d/%m/%Y %H:%M")
                if incidencia.fecha_cierre else ""
            ),
            incidencia.solucion or "",
        ])

    return response


@login_required
@pertenece_a_grupo("Administradores")
def exportar_incidencias_excel(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Incidencias"

    encabezados = [
        "ID",
        "Título",
        "Material",
        "Código material",
        "Prioridad",
        "Estado",
        "Usuario",
        "Fecha creación",
        "Fecha cierre",
        "Solución",
    ]

    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=texto)
        
    color_corporativo = "0051A0"

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color=color_corporativo,
            end_color=color_corporativo,
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    incidencias = Incidencia.objects.select_related(
        "material",
        "usuario"
    ).all()
    
    busqueda = request.GET.get("busqueda", "")
    prioridad = request.GET.get("prioridad", "")
    estado = request.GET.get("estado", "")

    if busqueda:
        incidencias = incidencias.filter(
            Q(titulo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if prioridad:
        incidencias = incidencias.filter(prioridad=prioridad)

    if estado:
        incidencias = incidencias.filter(estado=estado)

    fila = 2

    for incidencia in incidencias:
        hoja.cell(fila, 1, incidencia.id)
        hoja.cell(fila, 2, incidencia.titulo)
        hoja.cell(fila, 3, incidencia.material.nombre)
        hoja.cell(fila, 4, incidencia.material.codigo_inventario)
        hoja.cell(fila, 5, incidencia.get_prioridad_display())
        hoja.cell(fila, 6, incidencia.get_estado_display())

        if incidencia.usuario:
            hoja.cell(fila, 7, incidencia.usuario.username)
        else:
            hoja.cell(fila, 7, "")

        hoja.cell(
            fila,
            8,
            incidencia.fecha_creacion.strftime("%d/%m/%Y %H:%M")
        )

        if incidencia.fecha_cierre:
            hoja.cell(
                fila,
                9,
                incidencia.fecha_cierre.strftime("%d/%m/%Y %H:%M")
            )
        else:
            hoja.cell(fila, 9, "")

        hoja.cell(fila, 10, incidencia.solucion or "")

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="incidencias.xlsx"'
    
    for columna in hoja.columns:
        max_length = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        hoja.column_dimensions[letra_columna].width = max_length + 2

    hoja.auto_filter.ref = hoja.dimensions
    hoja.freeze_panes = "A2"

    workbook.save(response)

    return response

@login_required
@pertenece_a_grupo("Administradores")
def exportar_incidencias_pdf(request):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer)

    elementos = []
    estilos = getSampleStyleSheet()

    logo_path = settings.BASE_DIR / "static" / "images" / "logo_monlau.png"

    if logo_path.exists():
        elementos.append(Image(str(logo_path), width=90, height=55))

    elementos.append(Paragraph("<b>Informe de incidencias</b>", estilos["Title"]))
    elementos.append(Spacer(1, 12))

    datos = [
        ["ID", "Título", "Material", "Prioridad", "Estado"]
    ]

    incidencias = Incidencia.objects.select_related("material").all()

    busqueda = request.GET.get("busqueda", "")
    prioridad = request.GET.get("prioridad", "")
    estado = request.GET.get("estado", "")

    if busqueda:
        incidencias = incidencias.filter(
            Q(titulo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
            )

    if prioridad:
            incidencias = incidencias.filter(prioridad=prioridad)

    if estado:
        incidencias = incidencias.filter(estado=estado)

    for incidencia in incidencias:
        datos.append([
            incidencia.id,
            incidencia.titulo,
            incidencia.material.nombre,
            incidencia.get_prioridad_display(),
            incidencia.get_estado_display(),
        ])

    tabla = Table(datos, repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051A0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="incidencias.pdf"'

    return response
