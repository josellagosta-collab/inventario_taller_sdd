from .forms import (
    PrestamoForm,
    LineaPrestamoForm,
    ReservaForm,
) 
from .models import Prestamo, LineaPrestamo, Reserva
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from inventario.models import MovimientoInventario
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.conf import settings
from django.db.models import Q
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

@login_required
def crear_prestamo(request):
    if request.method == "POST":
        prestamo_form = PrestamoForm(request.POST)
        linea_form = LineaPrestamoForm(request.POST)

        if prestamo_form.is_valid() and linea_form.is_valid():
            prestamo = prestamo_form.save()

            linea = linea_form.save(commit=False)
            linea.prestamo = prestamo
            linea.save()

            material = linea.material
            material.estado = "prestado"
            material.save()

            MovimientoInventario.objects.create(
                material=material,
                tipo="prestamo",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion=f"Préstamo registrado. Préstamo ID: {prestamo.id}"
        )
          

            return redirect("prestamos:lista_prestamos")

    else:
        prestamo_form = PrestamoForm()
        linea_form = LineaPrestamoForm()

    return render(request, "prestamos/crear_prestamo.html", {
        "prestamo_form": prestamo_form,
        "linea_form": linea_form,
    })
    
@login_required
def lista_prestamos(request):
    prestamos = Prestamo.objects.select_related(
        "usuario_receptor",
        "profesor_responsable"
    ).prefetch_related("lineas__material")

    usuarios = User.objects.all()

    estado = request.GET.get("estado", "")
    usuario_receptor = request.GET.get("usuario_receptor", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        prestamos = prestamos.filter(estado=estado)

    if usuario_receptor:
        prestamos = prestamos.filter(usuario_receptor_id=usuario_receptor)

    if profesor_responsable:
        prestamos = prestamos.filter(profesor_responsable_id=profesor_responsable)

    hoy = timezone.now().date()

    total_prestamos = Prestamo.objects.count()

    total_activos = Prestamo.objects.filter(
        estado="activo"
    ).count()

    total_devueltos = Prestamo.objects.filter(
        estado="devuelto"
    ).count()

    total_retrasados = Prestamo.objects.filter(
        estado="activo",
        fecha_prevista_devolucion__lt=hoy
    ).count()

    paginator = Paginator(prestamos, 10)
    numero_pagina = request.GET.get("page")
    pagina_prestamos = paginator.get_page(numero_pagina)

    return render(request, "prestamos/lista_prestamos.html", {
        "prestamos": pagina_prestamos,
        "usuarios": usuarios,
        "estado": estado,
        "usuario_receptor": usuario_receptor,
        "profesor_responsable": profesor_responsable,
        "estados": Prestamo.ESTADOS,
        "total_prestamos": total_prestamos,
        "total_activos": total_activos,
        "total_devueltos": total_devueltos,
        "total_retrasados": total_retrasados,
    })
    
@login_required
def devolver_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)

    if request.method == "POST":
        prestamo.estado = "devuelto"
        prestamo.fecha_devolucion_real = timezone.now().date()
        prestamo.save()

        for linea in prestamo.lineas.all():
            material = linea.material
            material.estado = "disponible"
            material.save()
            MovimientoInventario.objects.create(
                material=material,
                tipo="devolucion",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion=f"Devolución registrada. Préstamo ID: {prestamo.id}"
            )

        return redirect("prestamos:lista_prestamos")
    
        hoy = timezone.now().date()

    total_prestamos = Prestamo.objects.count()

    total_activos = Prestamo.objects.filter(
        estado="activo"
    ).count()

    total_devueltos = Prestamo.objects.filter(
        estado="devuelto"
    ).count()

    total_retrasados = Prestamo.objects.filter(
        estado="activo",
        fecha_prevista_devolucion__lt=hoy
    ).count()

    return render(request, "prestamos/devolver_prestamo.html", {
        "prestamo": prestamo,
        "total_prestamos": total_prestamos,
        "total_activos": total_activos,
        "total_devueltos": total_devueltos,
        "total_retrasados": total_retrasados,
    })
    
@login_required
def detalle_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(
        Prestamo.objects.select_related(
            "usuario_receptor",
            "profesor_responsable"
        ).prefetch_related("lineas__material"),
        id=prestamo_id
    )

    return render(request, "prestamos/detalle_prestamo.html", {
        "prestamo": prestamo,
    })
    
@login_required
def exportar_prestamos_excel(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Prestamos"

    encabezados = [
        "ID",
        "Usuario receptor",
        "Profesor responsable",
        "Materiales",
        "Fecha préstamo",
        "Fecha prevista devolución",
        "Fecha devolución real",
        "Estado",
    ]

    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=texto)
        
    color_corporativo = "0051A0"

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color=color_corporativo,
            end_color=color_corporativo,
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    prestamos = Prestamo.objects.select_related(
        "usuario_receptor",
        "profesor_responsable"
    ).prefetch_related("lineas__material")
    
    estado = request.GET.get("estado", "")
    usuario_receptor = request.GET.get("usuario_receptor", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        prestamos = prestamos.filter(estado=estado)

    if usuario_receptor:
        prestamos = prestamos.filter(usuario_receptor_id=usuario_receptor)

    if profesor_responsable:
        prestamos = prestamos.filter(profesor_responsable_id=profesor_responsable)

    fila = 2

    for prestamo in prestamos:
        materiales = ", ".join(
            [
                f"{linea.material.nombre} x {linea.cantidad}"
                for linea in prestamo.lineas.all()
            ]
        )

        hoja.cell(fila, 1, prestamo.id)
        hoja.cell(fila, 2, prestamo.usuario_receptor.username)
        hoja.cell(fila, 3, prestamo.profesor_responsable.username)
        hoja.cell(fila, 4, materiales)
        hoja.cell(fila, 5, prestamo.fecha_prestamo.strftime("%d/%m/%Y"))

        hoja.cell(
            fila,
            6,
            prestamo.fecha_prevista_devolucion.strftime("%d/%m/%Y")
        )

        if prestamo.fecha_devolucion_real:
            hoja.cell(
                fila,
                7,
                prestamo.fecha_devolucion_real.strftime("%d/%m/%Y")
            )
        else:
            hoja.cell(fila, 7, "")
        hoja.cell(fila, 8, prestamo.get_estado_display())

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="prestamos.xlsx"'
    
    for columna in hoja.columns:
        max_length = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))

        hoja.column_dimensions[letra_columna].width = max_length + 2

    hoja.auto_filter.ref = hoja.dimensions
    hoja.freeze_panes = "A2"

    workbook.save(response)

    return response

@login_required
def exportar_prestamos_pdf(request):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer)

    elementos = []
    estilos = getSampleStyleSheet()

    logo_path = settings.BASE_DIR / "static" / "images" / "logo_monlau.png"

    if logo_path.exists():
        elementos.append(Image(str(logo_path), width=90, height=55))

    elementos.append(Paragraph("<b>Informe de préstamos</b>", estilos["Title"]))
    elementos.append(Spacer(1, 12))

    datos = [
        ["ID", "Usuario", "Profesor", "Material", "Fecha", "Prevista", "Estado"]
    ]

    prestamos = Prestamo.objects.select_related(
        "usuario_receptor",
        "profesor_responsable"
    ).prefetch_related("lineas__material")

    estado = request.GET.get("estado", "")
    usuario_receptor = request.GET.get("usuario_receptor", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        prestamos = prestamos.filter(estado=estado)

    if usuario_receptor:
        prestamos = prestamos.filter(usuario_receptor_id=usuario_receptor)

    if profesor_responsable:
        prestamos = prestamos.filter(profesor_responsable_id=profesor_responsable)

    for prestamo in prestamos:
        materiales = ", ".join(
            [
                f"{linea.material.nombre} x {linea.cantidad}"
                for linea in prestamo.lineas.all()
            ]
        )

        datos.append([
            prestamo.id,
            prestamo.usuario_receptor.username,
            prestamo.profesor_responsable.username,
            materiales,
            prestamo.fecha_prestamo.strftime("%d/%m/%Y"),
            prestamo.fecha_prevista_devolucion.strftime("%d/%m/%Y"),
            prestamo.get_estado_display(),
        ])

    tabla = Table(datos, repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051A0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="prestamos.pdf"'

    return response

@login_required
def crear_reserva(request):
    if request.method == "POST":
        form = ReservaForm(request.POST)

        if form.is_valid():
            reserva = form.save()

            material = reserva.material
            material.estado = "reservado"
            material.save()

            MovimientoInventario.objects.create(
                material=material,
                tipo="edicion",
                usuario=request.user if request.user.is_authenticated else None,
                descripcion=f"Reserva creada. Reserva ID: {reserva.id}"
            )

            return redirect("prestamos:lista_reservas")

    else:
        form = ReservaForm()

    return render(request, "prestamos/crear_reserva.html", {
        "form": form,
    })
    
@login_required
def lista_reservas(request):
    reservas = Reserva.objects.select_related(
        "usuario_reserva",
        "profesor_responsable",
        "material"
    ).all()

    usuarios = User.objects.all()

    estado = request.GET.get("estado", "")
    busqueda = request.GET.get("busqueda", "")
    usuario_reserva = request.GET.get("usuario_reserva", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        reservas = reservas.filter(estado=estado)

    if busqueda:
        reservas = reservas.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if usuario_reserva:
        reservas = reservas.filter(usuario_reserva_id=usuario_reserva)

    if profesor_responsable:
        reservas = reservas.filter(profesor_responsable_id=profesor_responsable)
      
    paginator = Paginator(reservas, 10)
    numero_pagina = request.GET.get("page")
    pagina_reservas = paginator.get_page(numero_pagina)  

    return render(request, "prestamos/lista_reservas.html", {
        "reservas": pagina_reservas,
        "usuarios": usuarios,
        "estados": Reserva.ESTADOS,
        "estado": estado,
        "busqueda": busqueda,
        "usuario_reserva": usuario_reserva,
        "profesor_responsable": profesor_responsable,
    })
    
@login_required
def cancelar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)

    if request.method == "POST":
        reserva.estado = "cancelada"
        reserva.save()

        material = reserva.material
        material.estado = "disponible"
        material.save()

        MovimientoInventario.objects.create(
            material=material,
            tipo="edicion",
            usuario=request.user if request.user.is_authenticated else None,
            descripcion=f"Reserva cancelada. Reserva ID: {reserva.id}"
        )

        return redirect("prestamos:lista_reservas")

    return render(request, "prestamos/cancelar_reserva.html", {
        "reserva": reserva,
    })
    
@login_required
def convertir_reserva_en_prestamo(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)

    if request.method == "POST":
        prestamo = Prestamo.objects.create(
            usuario_receptor=reserva.usuario_reserva,
            profesor_responsable=reserva.profesor_responsable,
            fecha_prevista_devolucion=reserva.fecha_prevista_recogida,
            estado="activo",
            observaciones=f"Préstamo generado desde reserva ID {reserva.id}"
        )

        LineaPrestamo.objects.create(
            prestamo=prestamo,
            material=reserva.material,
            cantidad=reserva.cantidad
        )

        reserva.estado = "convertida"
        reserva.save()

        material = reserva.material
        material.estado = "prestado"
        material.save()

        MovimientoInventario.objects.create(
            material=material,
            tipo="prestamo",
            usuario=request.user if request.user.is_authenticated else None,
            descripcion=f"Reserva convertida en préstamo. Reserva ID: {reserva.id}. Préstamo ID: {prestamo.id}"
        )

        return redirect("prestamos:detalle_prestamo", prestamo_id=prestamo.id)

    return render(request, "prestamos/convertir_reserva.html", {
        "reserva": reserva,
    })
    
@login_required
def exportar_reservas_excel(request):
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Reservas"

    encabezados = [
        "ID",
        "Usuario reserva",
        "Profesor responsable",
        "Material",
        "Código material",
        "Cantidad",
        "Fecha reserva",
        "Fecha prevista recogida",
        "Estado",
        "Observaciones",
    ]

    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=texto)
        
    color_corporativo = "0051A0"

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color=color_corporativo,
            end_color=color_corporativo,
            fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center")

    reservas = Reserva.objects.select_related(
        "usuario_reserva",
        "profesor_responsable",
        "material"
    ).all()
    
    estado = request.GET.get("estado", "")
    busqueda = request.GET.get("busqueda", "")
    usuario_reserva = request.GET.get("usuario_reserva", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        reservas = reservas.filter(estado=estado)

    if busqueda:
        reservas = reservas.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if usuario_reserva:
        reservas = reservas.filter(usuario_reserva_id=usuario_reserva)

    if profesor_responsable:
        reservas = reservas.filter(profesor_responsable_id=profesor_responsable)

    fila = 2

    for reserva in reservas:
        hoja.cell(fila, 1, reserva.id)
        hoja.cell(fila, 2, reserva.usuario_reserva.username)
        hoja.cell(fila, 3, reserva.profesor_responsable.username)
        hoja.cell(fila, 4, reserva.material.nombre)
        hoja.cell(fila, 5, reserva.material.codigo_inventario)
        hoja.cell(fila, 6, reserva.cantidad)
        hoja.cell(fila, 7, reserva.fecha_reserva.strftime("%d/%m/%Y"))
        hoja.cell(fila, 8, reserva.fecha_prevista_recogida.strftime("%d/%m/%Y"))
        hoja.cell(fila, 9, reserva.get_estado_display())
        hoja.cell(fila, 10, reserva.observaciones or "")

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="reservas.xlsx"'
    
    for columna in hoja.columns:
        max_length = 0
        letra_columna = get_column_letter(columna[0].column)

    for celda in columna:
        if celda.value:
            max_length = max(max_length, len(str(celda.value)))

    hoja.column_dimensions[letra_columna].width = max_length + 2

    hoja.auto_filter.ref = hoja.dimensions
    hoja.freeze_panes = "A2"

    workbook.save(response)

    return response

@login_required
def exportar_reservas_pdf(request):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer)

    elementos = []
    estilos = getSampleStyleSheet()

    logo_path = settings.BASE_DIR / "static" / "images" / "logo_monlau.png"

    if logo_path.exists():
        elementos.append(Image(str(logo_path), width=90, height=55))

    elementos.append(Paragraph("<b>Informe de reservas</b>", estilos["Title"]))
    elementos.append(Spacer(1, 12))

    datos = [
        ["ID", "Usuario", "Profesor", "Material", "Cantidad", "Recogida", "Estado"]
    ]

    reservas = Reserva.objects.select_related(
        "usuario_reserva",
        "profesor_responsable",
        "material"
    ).all()

    estado = request.GET.get("estado", "")
    busqueda = request.GET.get("busqueda", "")
    usuario_reserva = request.GET.get("usuario_reserva", "")
    profesor_responsable = request.GET.get("profesor_responsable", "")

    if estado:
        reservas = reservas.filter(estado=estado)

    if busqueda:
        reservas = reservas.filter(
            Q(material__nombre__icontains=busqueda) |
            Q(material__codigo_inventario__icontains=busqueda)
        )

    if usuario_reserva:
        reservas = reservas.filter(usuario_reserva_id=usuario_reserva)

    if profesor_responsable:
        reservas = reservas.filter(profesor_responsable_id=profesor_responsable)

    for reserva in reservas:
        datos.append([
            reserva.id,
            reserva.usuario_reserva.username,
            reserva.profesor_responsable.username,
            reserva.material.nombre,
            reserva.cantidad,
            reserva.fecha_prevista_recogida.strftime("%d/%m/%Y"),
            reserva.get_estado_display(),
        ])

    tabla = Table(datos, repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051A0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elementos.append(tabla)
    pdf.build(elementos)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reservas.pdf"'

    return response

@login_required
def detalle_reserva(request, reserva_id):
    reserva = get_object_or_404(
        Reserva.objects.select_related(
            "usuario_reserva",
            "profesor_responsable",
            "material"
        ),
        id=reserva_id
    )

    return render(request, "prestamos/detalle_reserva.html", {
        "reserva": reserva,
    })
    
@login_required
def actualizar_reservas_caducadas(request):
    hoy = timezone.now().date()

    reservas = Reserva.objects.filter(
        estado="activa",
        fecha_prevista_recogida__lt=hoy
    )

    for reserva in reservas:
        reserva.estado = "caducada"
        reserva.save()

        material = reserva.material
        material.estado = "disponible"
        material.save()

        MovimientoInventario.objects.create(
            material=material,
            tipo="edicion",
            usuario=request.user if request.user.is_authenticated else None,
            descripcion=f"Reserva caducada automáticamente. Reserva ID: {reserva.id}"
        )

    return redirect("prestamos:lista_reservas")